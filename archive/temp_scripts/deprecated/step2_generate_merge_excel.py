#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
步骤2：生成Person合并建议Excel（纯Python版）
"""
import pickle
from pathlib import Path
from collections import defaultdict, Counter
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

print("=" * 80)
print("生成Person合并建议Excel")
print("=" * 80)

# 1. 加载数据
print("\n[1/4] 加载数据...")
with open('person_database.pkl', 'rb') as f:
    db = pickle.load(f)

persons = db['persons']
person_index = db['person_index']
conversation_persons = db['conversation_persons']

print(f"  Person实例: {len(persons):,}")
print(f"  唯一人名: {len(person_index):,}")

# 2. 构建合并候选组
print("\n[2/4] 分析合并候选...")

merge_groups = []

def get_person_and_relation(name):
    """提取 "XXX的YYY" 中的XXX和YYY"""
    match = re.match(r'(.+)的(.+)', name)
    if match:
        return match.group(1), match.group(2)
    return None, name

# 策略A: 按"XXX的YYY"模式分组
print("  策略A: 明确关系词...")
relation_groups = defaultdict(lambda: defaultdict(list))

for name in person_index.keys():
    person, relation = get_person_and_relation(name)
    if person:
        relation_groups[person][relation].append(name)

# 生成明确关系词合并组
count_a = 0
for person_name, relations in relation_groups.items():
    for relation_word, names in relations.items():
        if len(names) <= 1:
            continue

        variants = []
        for name in names:
            instances = person_index[name]
            convs = Counter(persons[idx]['conversation'] for idx in instances)
            aliases = set()
            for idx in instances:
                if persons[idx]['aliases']:
                    aliases.update(persons[idx]['aliases'])

            variants.append({
                'name': name,
                'count': len(instances),
                'conversations': list(convs.keys())[:10],
                'conv_count': len(convs),
                'aliases': list(aliases)[:10]
            })

        merge_groups.append({
            'suggested_name': f"{person_name}的{relation_word}",
            'confidence': 'high',
            'reason': f'都明确指向{person_name}的{relation_word}',
            'variants': variants,
            'total_instances': sum(v['count'] for v in variants)
        })
        count_a += 1

print(f"    找到{count_a}组")

# 策略B: 同一对话内的关系词匹配
print("  策略B: 同对话内匹配...")
count_b = 0
processed_pairs = set()

for conv_name, person_names in conversation_persons.items():
    person_names = list(person_names)

    for i, name1 in enumerate(person_names):
        person1, rel1 = get_person_and_relation(name1)

        for name2 in person_names[i+1:]:
            person2, rel2 = get_person_and_relation(name2)

            # 避免重复
            pair_key = tuple(sorted([name1, name2]))
            if pair_key in processed_pairs:
                continue

            should_merge = False
            reason = ""

            # 情况1: "XXX的YYY" 和 "YYY"
            if person1 and not person2 and rel1 == name2:
                should_merge = True
                reason = f'同对话"{conv_name}"内，"{name1}"和"{name2}"应为同一人'
            elif person2 and not person1 and rel2 == name1:
                should_merge = True
                reason = f'同对话"{conv_name}"内，"{name2}"和"{name1}"应为同一人'

            if should_merge:
                processed_pairs.add(pair_key)

                # 检查是否已在其他组
                already_in_group = False
                for group in merge_groups:
                    existing = [v['name'] for v in group['variants']]
                    if name1 in existing or name2 in existing:
                        already_in_group = True
                        break

                if not already_in_group:
                    variants = []
                    for name in [name1, name2]:
                        instances = person_index[name]
                        convs = Counter(persons[idx]['conversation'] for idx in instances)
                        aliases = set()
                        for idx in instances:
                            if persons[idx]['aliases']:
                                aliases.update(persons[idx]['aliases'])

                        variants.append({
                            'name': name,
                            'count': len(instances),
                            'conversations': list(convs.keys())[:10],
                            'conv_count': len(convs),
                            'aliases': list(aliases)[:10]
                        })

                    merge_groups.append({
                        'suggested_name': variants[0]['name'] if variants[0]['count'] >= variants[1]['count'] else variants[1]['name'],
                        'confidence': 'high',
                        'reason': reason,
                        'variants': variants,
                        'total_instances': sum(v['count'] for v in variants)
                    })
                    count_b += 1

print(f"    找到{count_b}组")

print(f"\n  合并建议总数: {len(merge_groups)}")

# 3. 生成Excel
print("\n[3/4] 生成Excel...")

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "合并建议"

# 表头
headers = ['ID', '置信度', '建议合并为', '变体详情', '总出现次数', '合并原因', '您的决定']
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 排序
confidence_order = {'high': 0, 'medium': 1, 'low': 2}
merge_groups.sort(key=lambda x: (confidence_order.get(x['confidence'], 3), -x['total_instances']))

# 添加数据
for idx, group in enumerate(merge_groups, 1):
    # 准备变体详情
    variants_text = []
    for v in group['variants']:
        conv_str = ', '.join(v['conversations'][:3])
        if v['conv_count'] > 3:
            conv_str += f" (共{v['conv_count']}个)"

        alias_str = ', '.join(v['aliases'][:3]) if v['aliases'] else '无'
        if len(v['aliases']) > 3:
            alias_str += f" +{len(v['aliases'])-3}个"

        variants_text.append(
            f"【{v['name']}】\n"
            f"  次数: {v['count']}\n"
            f"  对话: {conv_str}\n"
            f"  别名: {alias_str}"
        )

    row_data = [
        idx,
        group['confidence'],
        group['suggested_name'],
        '\n\n'.join(variants_text),
        group['total_instances'],
        group['reason'],
        ''  # 用户决定
    ]

    ws.append(row_data)

    # 设置单元格样式
    row_num = idx + 1
    for col_num in range(1, 8):
        cell = ws.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 置信度颜色
    conf_cell = ws.cell(row=row_num, column=2)
    if group['confidence'] == 'high':
        conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif group['confidence'] == 'medium':
        conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# 设置列宽
ws.column_dimensions['A'].width = 8   # ID
ws.column_dimensions['B'].width = 12  # 置信度
ws.column_dimensions['C'].width = 30  # 建议合并为
ws.column_dimensions['D'].width = 80  # 变体详情
ws.column_dimensions['E'].width = 12  # 总出现次数
ws.column_dimensions['F'].width = 50  # 合并原因
ws.column_dimensions['G'].width = 15  # 您的决定

# 冻结首行
ws.freeze_panes = 'A2'

# 保存
output_file = 'person_merge_suggestions.xlsx'
wb.save(output_file)

print(f"\n[4/4] 完成！")
print(f"  合并建议总数: {len(merge_groups)}")
high_count = sum(1 for g in merge_groups if g['confidence'] == 'high')
medium_count = sum(1 for g in merge_groups if g['confidence'] == 'medium')
low_count = sum(1 for g in merge_groups if g['confidence'] == 'low')
print(f"    高置信度: {high_count}")
print(f"    中等置信度: {medium_count}")
print(f"    低置信度: {low_count}")
print(f"\n✅ Excel文件: {output_file}")
print(f"\n📝 使用说明:")
print(f"  1. 打开 {output_file}")
print(f"  2. 审核每条合并建议")
print(f"  3. 在'您的决定'列填写:")
print(f"     - approve (同意合并)")
print(f"     - reject (拒绝)")
print(f"  4. 保存后运行步骤3执行合并")
print("=" * 80)
